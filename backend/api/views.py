from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.http import HttpResponse
from .models import Direction, ManagerProfile, CompanyProfile, Department, Employee, Leave, Attendance, PasswordRecord
from .serializers import (
    DirectionSerializer, PasswordRecordSerializer, DepartmentSerializer, EmployeeSerializer,
    LeaveSerializer, AttendanceSerializer,
    RegisterSerializer, UserSerializer
)


class DirectionViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet pour les directions (lecture seule)"""
    queryset = Direction.objects.all()
    serializer_class = DirectionSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None


class PasswordRecordViewSet(viewsets.ModelViewSet):
    """ViewSet pour la table des mots de passe (admin et managers uniquement)"""
    queryset = PasswordRecord.objects.all()
    serializer_class = PasswordRecordSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ['user__first_name', 'user__last_name', 'user__username', 'role']
    ordering_fields = ['user__last_name', 'role', 'created_at']

    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return PasswordRecord.objects.all()
        # Entreprise : ne voit que les mots de passe de ses employés
        try:
            company = user.company_profile
            employee_users = Employee.objects.filter(
                department=company.department
            ).values_list('user', flat=True)
            return PasswordRecord.objects.filter(user__in=employee_users)
        except CompanyProfile.DoesNotExist:
            pass
        # Manager : ne voit que les mots de passe de ses directions
        if user.is_staff:
            try:
                profile = user.manager_profile
                direction_names = list(profile.directions.values_list('name', flat=True))
                if direction_names:
                    employee_users = Employee.objects.filter(
                        direction__in=direction_names
                    ).values_list('user', flat=True)
                    return PasswordRecord.objects.filter(user__in=employee_users)
            except ManagerProfile.DoesNotExist:
                pass
            return PasswordRecord.objects.all()
        return PasswordRecord.objects.none()


class DepartmentViewSet(viewsets.ModelViewSet):
    """ViewSet pour les entreprises"""
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return Department.objects.all()
        # Entreprise : ne voit que son entreprise
        try:
            company = user.company_profile
            return Department.objects.filter(id=company.department.id)
        except CompanyProfile.DoesNotExist:
            pass
        # Manager : voit tous les entreprises
        if user.is_staff:
            return Department.objects.all()
        # Employé : voit son entreprise
        try:
            emp = user.employee_profile
            if emp.department:
                return Department.objects.filter(id=emp.department.id)
        except Employee.DoesNotExist:
            pass
        return Department.objects.none()

    @action(detail=True, methods=['get'])
    def employees(self, request, pk=None):
        """Récupère tous les employés d'une entreprise"""
        department = self.get_object()
        employees = department.employees.all()
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)


class EmployeeViewSet(viewsets.ModelViewSet):
    """ViewSet pour les employés"""
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None
    filterset_fields = ['department', 'status', 'position']
    search_fields = ['first_name', 'last_name', 'email', 'position']
    ordering_fields = ['created_at', 'hire_date', 'last_name']

    def get_queryset(self):
        """Filtre les employés selon le rôle de l'utilisateur"""
        user = self.request.user
        queryset = Employee.objects.all()

        # Admin (superuser) voit tout
        if user.is_superuser:
            return queryset

        # Entreprise : ne voit que les employés de son entreprise
        if user.is_staff and not user.is_superuser:
            try:
                company = user.company_profile
                return queryset.filter(department=company.department)
            except CompanyProfile.DoesNotExist:
                pass

            # Manager : ne voit que les employés de ses directions
            try:
                profile = user.manager_profile
                direction_names = list(profile.directions.values_list('name', flat=True))
                if direction_names:
                    return queryset.filter(direction__in=direction_names)
                return queryset.none()
            except ManagerProfile.DoesNotExist:
                return queryset.none()

        # Employé : ne voit que son propre profil
        return queryset.filter(user=user)

    @action(detail=False, methods=['get'])
    def by_department(self, request):
        """Récupère les employés groupés par entreprise"""
        departments = Department.objects.all()
        employee_qs = self.get_queryset()
        result = []
        for dept in departments:
            dept_employees = employee_qs.filter(department=dept)
            if dept_employees.exists():
                result.append({
                    'department': DepartmentSerializer(dept).data,
                    'employees': EmployeeSerializer(dept_employees, many=True).data
                })
        return Response(result)


class LeaveViewSet(viewsets.ModelViewSet):
    """ViewSet pour les congés"""
    queryset = Leave.objects.all()
    pagination_class = None
    serializer_class = LeaveSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['employee', 'status', 'leave_type']
    search_fields = ['employee__first_name', 'employee__last_name', 'reason']
    ordering_fields = ['created_at', 'start_date']

    def get_queryset(self):
        """Filtre les congés selon le rôle de l'utilisateur"""
        user = self.request.user
        queryset = Leave.objects.all()

        if user.is_superuser:
            return queryset

        if user.is_staff and not user.is_superuser:
            try:
                company = user.company_profile
                return queryset.filter(employee__department=company.department)
            except CompanyProfile.DoesNotExist:
                pass

            try:
                profile = user.manager_profile
                direction_names = list(profile.directions.values_list('name', flat=True))
                if direction_names:
                    return queryset.filter(employee__direction__in=direction_names)
                return queryset.none()
            except ManagerProfile.DoesNotExist:
                return queryset.none()

        return queryset.filter(employee__user=user)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approuve une demande de congé (validation double : Manager puis Entreprise)"""
        leave = self.get_object()
        user = request.user

        if user.is_superuser:
            # Admin : approbation directe
            leave.status = 'approved'
            leave.approved_by = user
            if not leave.manager_approved_by:
                leave.manager_approved_by = user
        else:
            # Vérifier si c'est un responsable entreprise
            is_entreprise = False
            try:
                user.company_profile
                is_entreprise = True
            except CompanyProfile.DoesNotExist:
                pass

            if is_entreprise:
                # Entreprise : ne peut valider que si le manager a déjà validé
                if leave.status != 'manager_approved':
                    return Response(
                        {"error": "Ce congé doit d'abord être validé par le manager de direction."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                leave.status = 'approved'
                leave.approved_by = user
            elif user.is_staff:
                # Manager : première validation
                if leave.status != 'pending':
                    return Response(
                        {"error": "Ce congé n'est plus en attente de validation manager."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                leave.status = 'manager_approved'
                leave.manager_approved_by = user

        leave.save()
        serializer = self.get_serializer(leave)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Rejette une demande de congé (possible à n'importe quelle étape)"""
        leave = self.get_object()
        if leave.status in ('approved', 'rejected'):
            return Response(
                {"error": "Ce congé a déjà été traité."},
                status=status.HTTP_400_BAD_REQUEST
            )
        leave.status = 'rejected'
        leave.approved_by = request.user
        leave.save()
        serializer = self.get_serializer(leave)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Récupère toutes les demandes en attente (pending + manager_approved)"""
        pending_leaves = self.get_queryset().filter(status__in=['pending', 'manager_approved'])
        serializer = self.get_serializer(pending_leaves, many=True)
        return Response(serializer.data)


class AttendanceViewSet(viewsets.ModelViewSet):
    pagination_class = None
    """ViewSet pour les présences"""
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['employee', 'status', 'date']
    search_fields = ['employee__first_name', 'employee__last_name']
    ordering_fields = ['date']

    def get_queryset(self):
        """Filtre les présences selon le rôle de l'utilisateur"""
        user = self.request.user
        queryset = Attendance.objects.all()

        if user.is_superuser:
            return queryset

        if user.is_staff and not user.is_superuser:
            try:
                company = user.company_profile
                return queryset.filter(employee__department=company.department)
            except CompanyProfile.DoesNotExist:
                pass

            try:
                profile = user.manager_profile
                direction_names = list(profile.directions.values_list('name', flat=True))
                if direction_names:
                    return queryset.filter(employee__direction__in=direction_names)
                return queryset.none()
            except ManagerProfile.DoesNotExist:
                return queryset.none()

        return queryset.filter(employee__user=user)

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Récupère les présences d'aujourd'hui"""
        from datetime import date
        today_attendance = self.queryset.filter(date=date.today())
        serializer = self.get_serializer(today_attendance, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        """Récupère les présences par employé"""
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response(
                {"error": "employee_id est requis"},
                status=status.HTTP_400_BAD_REQUEST
            )

        attendances = self.queryset.filter(employee_id=employee_id)
        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)


class RegisterView(APIView):
    """Vue pour l'enregistrement d'utilisateurs"""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            refresh = RefreshToken.for_user(user)
            return Response({
                'user': UserSerializer(user).data,
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    """Vue pour la connexion personnalisée"""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        user = authenticate(username=username, password=password)

        if user is not None:
            # Déterminer automatiquement le rôle basé sur les permissions
            if user.is_superuser:
                role = 'admin'
            else:
                try:
                    user.company_profile
                    role = 'entreprise'
                except CompanyProfile.DoesNotExist:
                    if user.is_staff:
                        role = 'manager'
                    else:
                        role = 'employee'

            # Récupérer les directions du manager
            managed_directions = []
            managed_department = None
            if role == 'manager':
                try:
                    profile = user.manager_profile
                    managed_directions = list(profile.directions.values_list('name', flat=True))
                except ManagerProfile.DoesNotExist:
                    pass
            elif role == 'entreprise':
                try:
                    managed_department = {
                        'id': user.company_profile.department.id,
                        'name': user.company_profile.department.name,
                    }
                except CompanyProfile.DoesNotExist:
                    pass

            refresh = RefreshToken.for_user(user)
            return Response({
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'name': user.get_full_name() or user.username,
                    'role': role,
                    'managed_directions': managed_directions,
                    'managed_department': managed_department,
                },
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        else:
            return Response(
                {"error": "Identifiants invalides"},
                status=status.HTTP_401_UNAUTHORIZED
            )


class ChangePasswordView(APIView):
    """Vue pour changer le mot de passe"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')

        if not old_password or not new_password:
            return Response(
                {"error": "Ancien et nouveau mot de passe requis"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérifier l'ancien mot de passe
        if not user.check_password(old_password):
            return Response(
                {"error": "Ancien mot de passe incorrect"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Vérifier la longueur du nouveau mot de passe
        if len(new_password) < 6:
            return Response(
                {"error": "Le nouveau mot de passe doit contenir au moins 6 caractères"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Changer le mot de passe
        user.set_password(new_password)
        user.save()

        return Response({"message": "Mot de passe modifié avec succès"})


class DashboardStatsView(APIView):
    """Vue pour les statistiques du tableau de bord"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from datetime import date

        user = request.user
        employees_qs = Employee.objects.all()
        leaves_qs = Leave.objects.all()
        attendance_qs = Attendance.objects.all()

        # Filtrer selon le rôle
        if not user.is_superuser and user.is_staff:
            try:
                company = user.company_profile
                employees_qs = employees_qs.filter(department=company.department)
                leaves_qs = leaves_qs.filter(employee__department=company.department)
                attendance_qs = attendance_qs.filter(employee__department=company.department)
            except CompanyProfile.DoesNotExist:
                try:
                    profile = user.manager_profile
                    direction_names = list(profile.directions.values_list('name', flat=True))
                    if direction_names:
                        employees_qs = employees_qs.filter(direction__in=direction_names)
                        leaves_qs = leaves_qs.filter(employee__direction__in=direction_names)
                        attendance_qs = attendance_qs.filter(employee__direction__in=direction_names)
                    else:
                        employees_qs = employees_qs.none()
                        leaves_qs = leaves_qs.none()
                        attendance_qs = attendance_qs.none()
                except ManagerProfile.DoesNotExist:
                    employees_qs = employees_qs.none()
                    leaves_qs = leaves_qs.none()
                    attendance_qs = attendance_qs.none()

        total_employees = employees_qs.count()
        total_departments = Department.objects.count()
        present_today = attendance_qs.filter(
            date=date.today(),
            status='present'
        ).count()
        on_leave_today = leaves_qs.filter(
            start_date__lte=date.today(),
            end_date__gte=date.today(),
            status='approved'
        ).count()

        return Response({
            'total_employees': total_employees,
            'total_departments': total_departments,
            'present_today': present_today,
            'on_leave_today': on_leave_today,
        })


# ===========================
# Rapports Excel
# ===========================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import date


def _get_excel_styles():
    """Retourne les styles communs pour les rapports Excel"""
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(name='Arial', bold=True, size=14, color='2D3748')
    title_align = Alignment(horizontal='center', vertical='center')
    return {
        'header_font': header_font, 'header_fill': header_fill, 'header_align': header_align,
        'cell_font': cell_font, 'cell_align': cell_align, 'center_align': center_align,
        'thin_border': thin_border, 'title_font': title_font, 'title_align': title_align,
    }


def _write_headers(ws, headers, col_widths, row=3, styles=None):
    """Ecrit les en-tetes du tableau Excel"""
    if not styles:
        styles = _get_excel_styles()
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_align']
        cell.border = styles['thin_border']
        # Gestion des colonnes > Z
        if col_idx <= 26:
            ws.column_dimensions[chr(64 + col_idx)].width = width
        else:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 25


def _write_title(ws, title, col_count, styles=None):
    """Ecrit le titre du rapport"""
    if not styles:
        styles = _get_excel_styles()
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(col_count)
    ws.merge_cells(f'A1:{last_col}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = styles['title_font']
    title_cell.alignment = styles['title_align']
    ws.row_dimensions[1].height = 35
    # Date de generation
    ws.merge_cells(f'A2:{last_col}2')
    date_cell = ws['A2']
    date_cell.value = f"Généré le {date.today().strftime('%d/%m/%Y')}"
    date_cell.font = Font(name='Arial', size=9, italic=True, color='666666')
    date_cell.alignment = Alignment(horizontal='center')


def _filter_employees(user):
    """Filtre les employes selon le role de l'utilisateur"""
    qs = Employee.objects.all()
    if user.is_superuser:
        return qs
    if user.is_staff and not user.is_superuser:
        try:
            company = user.company_profile
            return qs.filter(department=company.department)
        except CompanyProfile.DoesNotExist:
            pass
        try:
            profile = user.manager_profile
            direction_names = list(profile.directions.values_list('name', flat=True))
            if direction_names:
                return qs.filter(direction__in=direction_names)
            return qs.none()
        except ManagerProfile.DoesNotExist:
            return qs.none()
    return qs.filter(user=user)


def _make_response(wb, filename):
    """Cree la HttpResponse avec le fichier Excel"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class AttendanceReportView(APIView):
    """Rapport de presence Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user)
        employee_ids = employees.values_list('id', flat=True)
        attendances = Attendance.objects.filter(employee_id__in=employee_ids).select_related('employee', 'employee__department').order_by('-date')

        wb = Workbook()
        ws = wb.active
        ws.title = "Présences"

        headers = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'DIRECTION', 'DATE', 'ARRIVÉE', 'DÉPART', 'HEURES', 'STATUT']
        col_widths = [6, 30, 20, 25, 14, 12, 12, 10, 14]

        _write_title(ws, 'RAPPORT DE PRÉSENCE', len(headers), styles)
        _write_headers(ws, headers, col_widths, row=3, styles=styles)

        status_labels = {'present': 'Présent', 'absent': 'Absent', 'late': 'En retard', 'half_day': 'Demi-journée'}
        status_fills = {
            'present': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
            'absent': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
            'late': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
        }

        row_num = 4
        for idx, att in enumerate(attendances, 1):
            emp = att.employee
            row_data = [
                idx,
                emp.full_name,
                emp.department.name if emp.department else '-',
                emp.direction or '-',
                att.date.strftime('%d/%m/%Y') if att.date else '-',
                att.check_in.strftime('%H:%M') if att.check_in else '-',
                att.check_out.strftime('%H:%M') if att.check_out else '-',
                str(att.hours_worked) if att.hours_worked else '-',
                status_labels.get(att.status, att.status),
            ]
            row_fill = status_fills.get(att.status, PatternFill())
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.fill = row_fill
                cell.alignment = styles['center_align'] if col_idx in (1, 5, 6, 7, 8, 9) else styles['cell_align']
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # Total
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL : {len(attendances)} enregistrements')
        total_cell.font = Font(name='Arial', bold=True, size=11)
        total_cell.alignment = styles['center_align']
        total_cell.border = styles['thin_border']

        return _make_response(wb, f'Rapport_Presences_{date.today().strftime("%Y%m%d")}.xlsx')


class LeavesReportView(APIView):
    """Rapport des conges Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user)
        employee_ids = employees.values_list('id', flat=True)
        leaves = Leave.objects.filter(employee_id__in=employee_ids).select_related(
            'employee', 'employee__department', 'manager_approved_by', 'approved_by'
        ).order_by('-created_at')

        wb = Workbook()
        ws = wb.active
        ws.title = "Congés"

        headers = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'DIRECTION', 'TYPE', 'DÉBUT', 'FIN', 'JOURS', 'STATUT', 'VALIDÉ PAR (MANAGER)', 'APPROUVÉ PAR (ENTREPRISE)']
        col_widths = [6, 30, 20, 25, 18, 14, 14, 8, 18, 25, 25]

        _write_title(ws, 'RAPPORT DES CONGÉS', len(headers), styles)
        _write_headers(ws, headers, col_widths, row=3, styles=styles)

        type_labels = {
            'annual': 'Congé annuel', 'sick': 'Maladie', 'maternity': 'Maternité',
            'paternity': 'Paternité', 'unpaid': 'Sans solde', 'other': 'Autre',
        }
        status_labels = {
            'pending': 'En attente', 'manager_approved': 'Validé Manager',
            'approved': 'Approuvé', 'rejected': 'Rejeté',
        }
        status_fills = {
            'pending': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
            'manager_approved': PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'),
            'approved': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
            'rejected': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
        }

        row_num = 4
        for idx, leave in enumerate(leaves, 1):
            emp = leave.employee
            mgr_name = leave.manager_approved_by.get_full_name() if leave.manager_approved_by else '-'
            app_name = leave.approved_by.get_full_name() if leave.approved_by else '-'
            row_data = [
                idx,
                emp.full_name,
                emp.department.name if emp.department else '-',
                emp.direction or '-',
                type_labels.get(leave.leave_type, leave.leave_type),
                leave.start_date.strftime('%d/%m/%Y') if leave.start_date else '-',
                leave.end_date.strftime('%d/%m/%Y') if leave.end_date else '-',
                leave.days_count,
                status_labels.get(leave.status, leave.status),
                mgr_name,
                app_name,
            ]
            row_fill = status_fills.get(leave.status, PatternFill())
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.fill = row_fill
                cell.alignment = styles['center_align'] if col_idx in (1, 6, 7, 8, 9) else styles['cell_align']
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL : {len(leaves)} demandes de congés')
        total_cell.font = Font(name='Arial', bold=True, size=11)
        total_cell.alignment = styles['center_align']
        total_cell.border = styles['thin_border']

        return _make_response(wb, f'Rapport_Conges_{date.today().strftime("%Y%m%d")}.xlsx')


class DepartmentsReportView(APIView):
    """Rapport par entreprise Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user).select_related('department').order_by('department__name', 'last_name', 'first_name')

        wb = Workbook()
        ws = wb.active
        ws.title = "Par Entreprise"

        headers = ['N°', 'NOM COMPLET', 'EMAIL', 'TÉLÉPHONE', 'DIRECTION', 'POSTE', 'DATE EMBAUCHE', 'SALAIRE', 'STATUT']
        col_widths = [6, 30, 28, 16, 25, 20, 14, 14, 12]

        _write_title(ws, 'RAPPORT PAR ENTREPRISE', len(headers), styles)
        _write_headers(ws, headers, col_widths, row=3, styles=styles)

        dept_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        dept_font = Font(name='Arial', bold=True, size=11, color='1565C0')
        status_labels = {'active': 'Actif', 'inactive': 'Inactif', 'on_leave': 'En congé'}

        row_num = 4
        current_dept = None
        dept_count = 0
        global_idx = 0

        for emp in employees:
            dept_name = emp.department.name if emp.department else 'Sans entreprise'

            # Nouvelle entreprise : ecrire le sous-titre
            if dept_name != current_dept:
                if current_dept is not None:
                    row_num += 1  # ligne vide entre entreprises
                current_dept = dept_name
                dept_count = employees.filter(department=emp.department).count() if emp.department else 0

                from openpyxl.utils import get_column_letter
                last_col = get_column_letter(len(headers))
                ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
                dept_cell = ws.cell(row=row_num, column=1, value=f'{dept_name} ({dept_count} agents)')
                dept_cell.font = dept_font
                dept_cell.fill = dept_fill
                dept_cell.alignment = Alignment(horizontal='left', vertical='center')
                dept_cell.border = styles['thin_border']
                ws.row_dimensions[row_num].height = 28
                row_num += 1

            global_idx += 1
            row_data = [
                global_idx,
                emp.full_name,
                emp.email or '-',
                emp.phone or '-',
                emp.direction or '-',
                emp.position or '-',
                emp.hire_date.strftime('%d/%m/%Y') if emp.hire_date else '-',
                f'{emp.salary:,.0f}' if emp.salary else '-',
                status_labels.get(emp.status, emp.status),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 7, 8, 9) else styles['cell_align']
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # Total general
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL GÉNÉRAL : {global_idx} employés')
        total_cell.font = Font(name='Arial', bold=True, size=11)
        total_cell.alignment = styles['center_align']
        total_cell.border = styles['thin_border']

        return _make_response(wb, f'Rapport_Entreprises_{date.today().strftime("%Y%m%d")}.xlsx')


class CompleteReportView(APIView):
    """Rapport RH complet (3 feuilles)"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user).select_related('department').order_by('last_name', 'first_name')
        employee_ids = employees.values_list('id', flat=True)

        wb = Workbook()

        # ===== FEUILLE 1 : EMPLOYES =====
        ws1 = wb.active
        ws1.title = "Employés"

        h1 = ['N°', 'NOM', 'PRÉNOM', 'EMAIL', 'TÉLÉPHONE', 'ENTREPRISE', 'DIRECTION', 'POSTE', 'MATRICULE', 'CNPS', 'DATE EMBAUCHE', 'SALAIRE', 'STATUT']
        w1 = [6, 20, 20, 28, 16, 20, 25, 20, 14, 14, 14, 14, 12]

        _write_title(ws1, 'RAPPORT RH COMPLET - EMPLOYÉS', len(h1), styles)
        _write_headers(ws1, h1, w1, row=3, styles=styles)

        status_labels = {'active': 'Actif', 'inactive': 'Inactif', 'on_leave': 'En congé'}
        row_num = 4
        for idx, emp in enumerate(employees, 1):
            row_data = [
                idx, emp.last_name, emp.first_name, emp.email or '-', emp.phone or '-',
                emp.department.name if emp.department else '-', emp.direction or '-',
                emp.position or '-', emp.matricule or '-', emp.cnps or '-',
                emp.hire_date.strftime('%d/%m/%Y') if emp.hire_date else '-',
                f'{emp.salary:,.0f}' if emp.salary else '-',
                status_labels.get(emp.status, emp.status),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 9, 10, 11, 12, 13) else styles['cell_align']
            ws1.row_dimensions[row_num].height = 20
            row_num += 1

        # ===== FEUILLE 2 : CONGES =====
        ws2 = wb.create_sheet("Congés")

        h2 = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'TYPE', 'DÉBUT', 'FIN', 'JOURS', 'STATUT', 'VALIDÉ MANAGER', 'APPROUVÉ ENTREPRISE']
        w2 = [6, 30, 20, 18, 14, 14, 8, 18, 25, 25]

        _write_title(ws2, 'RAPPORT RH COMPLET - CONGÉS', len(h2), styles)
        _write_headers(ws2, h2, w2, row=3, styles=styles)

        type_labels = {
            'annual': 'Congé annuel', 'sick': 'Maladie', 'maternity': 'Maternité',
            'paternity': 'Paternité', 'unpaid': 'Sans solde', 'other': 'Autre',
        }
        leave_status_labels = {
            'pending': 'En attente', 'manager_approved': 'Validé Manager',
            'approved': 'Approuvé', 'rejected': 'Rejeté',
        }

        leaves = Leave.objects.filter(employee_id__in=employee_ids).select_related(
            'employee', 'employee__department', 'manager_approved_by', 'approved_by'
        ).order_by('-created_at')

        row_num = 4
        for idx, leave in enumerate(leaves, 1):
            emp = leave.employee
            row_data = [
                idx, emp.full_name, emp.department.name if emp.department else '-',
                type_labels.get(leave.leave_type, leave.leave_type),
                leave.start_date.strftime('%d/%m/%Y') if leave.start_date else '-',
                leave.end_date.strftime('%d/%m/%Y') if leave.end_date else '-',
                leave.days_count,
                leave_status_labels.get(leave.status, leave.status),
                leave.manager_approved_by.get_full_name() if leave.manager_approved_by else '-',
                leave.approved_by.get_full_name() if leave.approved_by else '-',
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 5, 6, 7, 8) else styles['cell_align']
            ws2.row_dimensions[row_num].height = 20
            row_num += 1

        # ===== FEUILLE 3 : PRESENCES =====
        ws3 = wb.create_sheet("Présences")

        h3 = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'DATE', 'ARRIVÉE', 'DÉPART', 'HEURES', 'STATUT']
        w3 = [6, 30, 20, 14, 12, 12, 10, 14]

        _write_title(ws3, 'RAPPORT RH COMPLET - PRÉSENCES', len(h3), styles)
        _write_headers(ws3, h3, w3, row=3, styles=styles)

        att_status_labels = {'present': 'Présent', 'absent': 'Absent', 'late': 'En retard', 'half_day': 'Demi-journée'}
        attendances = Attendance.objects.filter(employee_id__in=employee_ids).select_related(
            'employee', 'employee__department'
        ).order_by('-date')

        row_num = 4
        for idx, att in enumerate(attendances, 1):
            emp = att.employee
            row_data = [
                idx, emp.full_name, emp.department.name if emp.department else '-',
                att.date.strftime('%d/%m/%Y') if att.date else '-',
                att.check_in.strftime('%H:%M') if att.check_in else '-',
                att.check_out.strftime('%H:%M') if att.check_out else '-',
                str(att.hours_worked) if att.hours_worked else '-',
                att_status_labels.get(att.status, att.status),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 4, 5, 6, 7, 8) else styles['cell_align']
            ws3.row_dimensions[row_num].height = 20
            row_num += 1

        return _make_response(wb, f'Rapport_RH_Complet_{date.today().strftime("%Y%m%d")}.xlsx')
