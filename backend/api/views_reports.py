from rest_framework import permissions
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import date

from .models import CompanyProfile, ManagerProfile, Department, Employee, Leave, Attendance


# ===========================
# Rapports Excel
# ===========================

def _get_excel_styles():
    """Retourne les styles communs pour les rapports Excel"""
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(name='Arial', bold=True, size=14, color='2D3748')
    title_align = Alignment(horizontal='center', vertical='center')
    return {
        'header_font': header_font, 'header_fill': header_fill, 'header_align': header_align,
        'cell_font': cell_font, 'cell_align': cell_align, 'center_align': center_align,
        'thin_border': thin_border, 'title_font': title_font, 'title_align': title_align,
    }


def _write_headers(ws, headers, col_widths, row=3, styles=None):
    """Ecrit les en-tetes du tableau Excel"""
    if not styles:
        styles = _get_excel_styles()
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_align']
        cell.border = styles['thin_border']
        # Gestion des colonnes > Z
        if col_idx <= 26:
            ws.column_dimensions[chr(64 + col_idx)].width = width
        else:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 25


def _write_title(ws, title, col_count, styles=None):
    """Ecrit le titre du rapport"""
    if not styles:
        styles = _get_excel_styles()
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(col_count)
    ws.merge_cells(f'A1:{last_col}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = styles['title_font']
    title_cell.alignment = styles['title_align']
    ws.row_dimensions[1].height = 35
    # Date de generation
    ws.merge_cells(f'A2:{last_col}2')
    date_cell = ws['A2']
    date_cell.value = f"Généré le {date.today().strftime('%d/%m/%Y')}"
    date_cell.font = Font(name='Arial', size=9, italic=True, color='666666')
    date_cell.alignment = Alignment(horizontal='center')


def _filter_employees(user):
    """Filtre les employes selon le role de l'utilisateur"""
    qs = Employee.objects.all()
    if user.is_superuser:
        return qs
    if user.is_staff and not user.is_superuser:
        try:
            company = user.company_profile
            return qs.filter(department=company.department)
        except CompanyProfile.DoesNotExist:
            pass
        try:
            profile = user.manager_profile
            direction_names = list(profile.directions.values_list('name', flat=True))
            if direction_names:
                return qs.filter(direction__in=direction_names)
            return qs.none()
        except ManagerProfile.DoesNotExist:
            return qs.none()
    return qs.filter(user=user)


def _make_response(wb, filename):
    """Cree la HttpResponse avec le fichier Excel"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class AttendanceReportView(APIView):
    """Rapport de presence Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user)
        employee_ids = employees.values_list('id', flat=True)
        attendances = Attendance.objects.filter(employee_id__in=employee_ids).select_related('employee', 'employee__department').order_by('-date')

        wb = Workbook()
        ws = wb.active
        ws.title = "Présences"

        headers = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'DIRECTION', 'DATE', 'ARRIVÉE', 'DÉPART', 'HEURES', 'STATUT']
        col_widths = [6, 30, 20, 25, 14, 12, 12, 10, 14]

        _write_title(ws, 'RAPPORT DE PRÉSENCE', len(headers), styles)
        _write_headers(ws, headers, col_widths, row=3, styles=styles)

        status_labels = {'present': 'Présent', 'absent': 'Absent', 'late': 'En retard', 'half_day': 'Demi-journée'}
        status_fills = {
            'present': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
            'absent': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
            'late': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
        }

        row_num = 4
        for idx, att in enumerate(attendances, 1):
            emp = att.employee
            row_data = [
                idx,
                emp.full_name,
                emp.department.name if emp.department else '-',
                emp.direction or '-',
                att.date.strftime('%d/%m/%Y') if att.date else '-',
                att.check_in.strftime('%H:%M') if att.check_in else '-',
                att.check_out.strftime('%H:%M') if att.check_out else '-',
                str(att.hours_worked) if att.hours_worked else '-',
                status_labels.get(att.status, att.status),
            ]
            row_fill = status_fills.get(att.status, PatternFill())
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.fill = row_fill
                cell.alignment = styles['center_align'] if col_idx in (1, 5, 6, 7, 8, 9) else styles['cell_align']
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # Total
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL : {len(attendances)} enregistrements')
        total_cell.font = Font(name='Arial', bold=True, size=11)
        total_cell.alignment = styles['center_align']
        total_cell.border = styles['thin_border']

        return _make_response(wb, f'Rapport_Presences_{date.today().strftime("%Y%m%d")}.xlsx')


class LeavesReportView(APIView):
    """Rapport des conges Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user)
        employee_ids = employees.values_list('id', flat=True)
        leaves = Leave.objects.filter(employee_id__in=employee_ids).select_related(
            'employee', 'employee__department', 'manager_approved_by', 'approved_by'
        ).order_by('-created_at')

        wb = Workbook()
        ws = wb.active
        ws.title = "Congés"

        headers = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'DIRECTION', 'TYPE', 'DÉBUT', 'FIN', 'JOURS', 'STATUT', 'VALIDÉ PAR (MANAGER)', 'APPROUVÉ PAR (ENTREPRISE)']
        col_widths = [6, 30, 20, 25, 18, 14, 14, 8, 18, 25, 25]

        _write_title(ws, 'RAPPORT DES CONGÉS', len(headers), styles)
        _write_headers(ws, headers, col_widths, row=3, styles=styles)

        type_labels = {
            'paid': 'Congé Payé', 'sick': 'Congé Maladie',
            'unpaid': 'Congé Sans Solde', 'parental': 'Congé Parental', 'other': 'Autre',
        }
        status_labels = {
            'pending': 'En attente', 'manager_approved': 'Validé Manager',
            'approved': 'Approuvé', 'rejected': 'Rejeté',
        }
        status_fills = {
            'pending': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
            'manager_approved': PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid'),
            'approved': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
            'rejected': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
        }

        row_num = 4
        for idx, leave in enumerate(leaves, 1):
            emp = leave.employee
            mgr_name = leave.manager_approved_by.get_full_name() if leave.manager_approved_by else '-'
            app_name = leave.approved_by.get_full_name() if leave.approved_by else '-'
            row_data = [
                idx,
                emp.full_name,
                emp.department.name if emp.department else '-',
                emp.direction or '-',
                type_labels.get(leave.leave_type, leave.leave_type),
                leave.start_date.strftime('%d/%m/%Y') if leave.start_date else '-',
                leave.end_date.strftime('%d/%m/%Y') if leave.end_date else '-',
                leave.days_count,
                status_labels.get(leave.status, leave.status),
                mgr_name,
                app_name,
            ]
            row_fill = status_fills.get(leave.status, PatternFill())
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.fill = row_fill
                cell.alignment = styles['center_align'] if col_idx in (1, 6, 7, 8, 9) else styles['cell_align']
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL : {len(leaves)} demandes de congés')
        total_cell.font = Font(name='Arial', bold=True, size=11)
        total_cell.alignment = styles['center_align']
        total_cell.border = styles['thin_border']

        return _make_response(wb, f'Rapport_Conges_{date.today().strftime("%Y%m%d")}.xlsx')


class DepartmentsReportView(APIView):
    """Rapport par entreprise Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user).select_related('department').order_by('department__name', 'last_name', 'first_name')

        wb = Workbook()
        ws = wb.active
        ws.title = "Par Entreprise"

        headers = ['N°', 'NOM COMPLET', 'EMAIL', 'TÉLÉPHONE', 'DIRECTION', 'POSTE', 'DATE EMBAUCHE', 'SALAIRE', 'STATUT']
        col_widths = [6, 30, 28, 16, 25, 20, 14, 14, 12]

        _write_title(ws, 'RAPPORT PAR ENTREPRISE', len(headers), styles)
        _write_headers(ws, headers, col_widths, row=3, styles=styles)

        dept_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        dept_font = Font(name='Arial', bold=True, size=11, color='1565C0')
        status_labels = {'active': 'Actif', 'inactive': 'Inactif', 'on_leave': 'En congé'}

        row_num = 4
        current_dept = None
        dept_count = 0
        global_idx = 0

        for emp in employees:
            dept_name = emp.department.name if emp.department else 'Sans entreprise'

            # Nouvelle entreprise : ecrire le sous-titre
            if dept_name != current_dept:
                if current_dept is not None:
                    row_num += 1  # ligne vide entre entreprises
                current_dept = dept_name
                dept_count = employees.filter(department=emp.department).count() if emp.department else 0

                from openpyxl.utils import get_column_letter
                last_col = get_column_letter(len(headers))
                ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
                dept_cell = ws.cell(row=row_num, column=1, value=f'{dept_name} ({dept_count} agents)')
                dept_cell.font = dept_font
                dept_cell.fill = dept_fill
                dept_cell.alignment = Alignment(horizontal='left', vertical='center')
                dept_cell.border = styles['thin_border']
                ws.row_dimensions[row_num].height = 28
                row_num += 1

            global_idx += 1
            row_data = [
                global_idx,
                emp.full_name,
                emp.email or '-',
                emp.phone or '-',
                emp.direction or '-',
                emp.position or '-',
                emp.hire_date.strftime('%d/%m/%Y') if emp.hire_date else '-',
                f'{emp.salary:,.0f}' if emp.salary else '-',
                status_labels.get(emp.status, emp.status),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 7, 8, 9) else styles['cell_align']
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # Total general
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(headers))
        ws.merge_cells(f'A{row_num}:{last_col}{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL GÉNÉRAL : {global_idx} employés')
        total_cell.font = Font(name='Arial', bold=True, size=11)
        total_cell.alignment = styles['center_align']
        total_cell.border = styles['thin_border']

        return _make_response(wb, f'Rapport_Entreprises_{date.today().strftime("%Y%m%d")}.xlsx')


class CompleteReportView(APIView):
    """Rapport RH complet (3 feuilles)"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        styles = _get_excel_styles()
        employees = _filter_employees(request.user).select_related('department').order_by('last_name', 'first_name')
        employee_ids = employees.values_list('id', flat=True)

        wb = Workbook()

        # ===== FEUILLE 1 : EMPLOYES =====
        ws1 = wb.active
        ws1.title = "Employés"

        h1 = ['N°', 'NOM', 'PRÉNOM', 'EMAIL', 'TÉLÉPHONE', 'ENTREPRISE', 'DIRECTION', 'POSTE', 'MATRICULE', 'CNPS', 'DATE EMBAUCHE', 'SALAIRE', 'STATUT']
        w1 = [6, 20, 20, 28, 16, 20, 25, 20, 14, 14, 14, 14, 12]

        _write_title(ws1, 'RAPPORT RH COMPLET - EMPLOYÉS', len(h1), styles)
        _write_headers(ws1, h1, w1, row=3, styles=styles)

        status_labels = {'active': 'Actif', 'inactive': 'Inactif', 'on_leave': 'En congé'}
        row_num = 4
        for idx, emp in enumerate(employees, 1):
            row_data = [
                idx, emp.last_name, emp.first_name, emp.email or '-', emp.phone or '-',
                emp.department.name if emp.department else '-', emp.direction or '-',
                emp.position or '-', emp.matricule or '-', emp.cnps or '-',
                emp.hire_date.strftime('%d/%m/%Y') if emp.hire_date else '-',
                f'{emp.salary:,.0f}' if emp.salary else '-',
                status_labels.get(emp.status, emp.status),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 9, 10, 11, 12, 13) else styles['cell_align']
            ws1.row_dimensions[row_num].height = 20
            row_num += 1

        # ===== FEUILLE 2 : CONGES =====
        ws2 = wb.create_sheet("Congés")

        h2 = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'TYPE', 'DÉBUT', 'FIN', 'JOURS', 'STATUT', 'VALIDÉ MANAGER', 'APPROUVÉ ENTREPRISE']
        w2 = [6, 30, 20, 18, 14, 14, 8, 18, 25, 25]

        _write_title(ws2, 'RAPPORT RH COMPLET - CONGÉS', len(h2), styles)
        _write_headers(ws2, h2, w2, row=3, styles=styles)

        type_labels = {
            'paid': 'Congé Payé', 'sick': 'Congé Maladie',
            'unpaid': 'Congé Sans Solde', 'parental': 'Congé Parental', 'other': 'Autre',
        }
        leave_status_labels = {
            'pending': 'En attente', 'manager_approved': 'Validé Manager',
            'approved': 'Approuvé', 'rejected': 'Rejeté',
        }

        leaves = Leave.objects.filter(employee_id__in=employee_ids).select_related(
            'employee', 'employee__department', 'manager_approved_by', 'approved_by'
        ).order_by('-created_at')

        row_num = 4
        for idx, leave in enumerate(leaves, 1):
            emp = leave.employee
            row_data = [
                idx, emp.full_name, emp.department.name if emp.department else '-',
                type_labels.get(leave.leave_type, leave.leave_type),
                leave.start_date.strftime('%d/%m/%Y') if leave.start_date else '-',
                leave.end_date.strftime('%d/%m/%Y') if leave.end_date else '-',
                leave.days_count,
                leave_status_labels.get(leave.status, leave.status),
                leave.manager_approved_by.get_full_name() if leave.manager_approved_by else '-',
                leave.approved_by.get_full_name() if leave.approved_by else '-',
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 5, 6, 7, 8) else styles['cell_align']
            ws2.row_dimensions[row_num].height = 20
            row_num += 1

        # ===== FEUILLE 3 : PRESENCES =====
        ws3 = wb.create_sheet("Présences")

        h3 = ['N°', 'NOM COMPLET', 'ENTREPRISE', 'DATE', 'ARRIVÉE', 'DÉPART', 'HEURES', 'STATUT']
        w3 = [6, 30, 20, 14, 12, 12, 10, 14]

        _write_title(ws3, 'RAPPORT RH COMPLET - PRÉSENCES', len(h3), styles)
        _write_headers(ws3, h3, w3, row=3, styles=styles)

        att_status_labels = {'present': 'Présent', 'absent': 'Absent', 'late': 'En retard', 'half_day': 'Demi-journée'}
        attendances = Attendance.objects.filter(employee_id__in=employee_ids).select_related(
            'employee', 'employee__department'
        ).order_by('-date')

        row_num = 4
        for idx, att in enumerate(attendances, 1):
            emp = att.employee
            row_data = [
                idx, emp.full_name, emp.department.name if emp.department else '-',
                att.date.strftime('%d/%m/%Y') if att.date else '-',
                att.check_in.strftime('%H:%M') if att.check_in else '-',
                att.check_out.strftime('%H:%M') if att.check_out else '-',
                str(att.hours_worked) if att.hours_worked else '-',
                att_status_labels.get(att.status, att.status),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles['cell_font']
                cell.border = styles['thin_border']
                cell.alignment = styles['center_align'] if col_idx in (1, 4, 5, 6, 7, 8) else styles['cell_align']
            ws3.row_dimensions[row_num].height = 20
            row_num += 1

        return _make_response(wb, f'Rapport_RH_Complet_{date.today().strftime("%Y%m%d")}.xlsx')
