"""
Script pour creer les comptes des 173 agents AZING IVOIR Sarl (Agents de bureau).
Entreprise: AZING IVOIR Sarl | Poste: Agent de bureau
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'empmanager.settings')
django.setup()

from django.contrib.auth.models import User
from api.models import Department, Employee, PasswordRecord

try:
    azing = Department.objects.get(name='AZING IVOIR Sarl')
    print(f"Entreprise trouvee : {azing.name} (ID: {azing.id})")
except Department.DoesNotExist:
    print("ERREUR : L'entreprise AZING IVOIR Sarl n'existe pas. Creation...")
    azing = Department.objects.create(
        name='AZING IVOIR Sarl',
        manager="N'GUESSAN DAMBO Benedicte",
        description='Prestation de service - Agents contractuels du Ministere de l Equipement et de l Entretien Routier'
    )
    print(f"Entreprise creee : {azing.name} (ID: {azing.id})")

agents = [
    {'nom': 'ABO', 'prenom': 'Brou Juliette', 'matricule': '3791-ME', 'service': 'DD Dabou'},
    {'nom': 'ABOH', 'prenom': 'Tano Francois', 'matricule': '3792-ME', 'service': 'DD Bongouanou'},
    {'nom': 'ACAFOU', 'prenom': 'Hyacinthe', 'matricule': '3793-ME', 'service': 'DR Alepe'},
    {'nom': 'ACHI', 'prenom': 'Chantal Georgette', 'matricule': '3794-ME', 'service': 'DD Agboville'},
    {'nom': 'ACHO', 'prenom': 'Apo Marie Chantal', 'matricule': '3795-ME', 'service': 'DD Tanda'},
    {'nom': 'ADOUKO', 'prenom': 'Koussi Firmin', 'matricule': '3796-ME', 'service': 'DD Abidjan'},
    {'nom': 'ADJOBI', 'prenom': 'Narcisse', 'matricule': '3797-ME', 'service': 'DD Soubre'},
    {'nom': 'AKE', 'prenom': "N'Cho Johness Eric", 'matricule': '3798-ME', 'service': 'DD Zuenoula'},
    {'nom': 'AMANI', 'prenom': 'Konan', 'matricule': '3800-ME', 'service': 'DAFP'},
    {'nom': 'AMARA', 'prenom': 'Ouattara', 'matricule': '3801-ME', 'service': 'DD Grand-Bassam'},
    {'nom': 'ASSOUMO', 'prenom': 'J. Eric', 'matricule': '3802-ME', 'service': 'DR Adzope'},
    {'nom': 'ATSE', 'prenom': 'Yapo Henri Didier', 'matricule': '3803-ME', 'service': 'DR Adzope'},
    {'nom': 'AVEREBY', 'prenom': 'Katie Phyrace', 'matricule': '3804-ME', 'service': 'DD Oume'},
    {'nom': 'BAMBA', 'prenom': 'Fessimata', 'matricule': '3805-ME', 'service': 'DR Man'},
    {'nom': 'BEDA', 'prenom': 'Nadine Micheline', 'matricule': '3807-ME', 'service': 'DR Adzope'},
    {'nom': 'BEHO', 'prenom': 'Djerou Agnes Epse BAH', 'matricule': '3808-ME', 'service': 'DD Sassandra'},
    {'nom': 'BLEURIEZ', 'prenom': 'Guiho Erwige Kevine Epse KOUASSI', 'matricule': '3809-ME', 'service': 'Inspection Generale'},
    {'nom': 'BLI', 'prenom': 'Fatou Patricia', 'matricule': '3810-ME', 'service': 'DR Korhogo'},
    {'nom': 'BOKA', 'prenom': 'Konan Anastasie', 'matricule': '3811-ME', 'service': 'DR Adzope'},
    {'nom': 'BONI', 'prenom': 'Assi Gildas Thibaut', 'matricule': '3812-ME', 'service': 'DD Issia'},
    {'nom': 'BOUCHOU', 'prenom': 'Placide Allali Didier Gilbert', 'matricule': '3813-ME', 'service': 'DD Tiassale'},
    {'nom': 'BOUSSOU', 'prenom': 'Armelie Rosine', 'matricule': '3814-ME', 'service': 'DR Agboville'},
    {'nom': 'BROU', 'prenom': 'Alla Koko Jeannette', 'matricule': '3815-ME', 'service': 'DD Dabou'},
    {'nom': 'BROU', 'prenom': "N'Guessan Noelle", 'matricule': '3816-ME', 'service': 'DD Touba'},
    {'nom': 'BROU', 'prenom': 'Obehi Chaye Melaine Annick', 'matricule': '3817-ME', 'service': 'DRH'},
    {'nom': 'CAMARA', 'prenom': 'Abdoulaye', 'matricule': '3818-ME', 'service': 'DAFP'},
    {'nom': 'CHIMENE', 'prenom': 'Nanou Eichem', 'matricule': '3819-ME', 'service': 'DR Aboisso'},
    {'nom': 'COULIBALY', 'prenom': 'Yopie Awa', 'matricule': '3820-ME', 'service': 'DR Adzope'},
    {'nom': 'DAGNOKO', 'prenom': 'Aminata', 'matricule': '3821-ME', 'service': 'DPE'},
    {'nom': 'DAGNOKO', 'prenom': 'Mariam', 'matricule': '3822-ME', 'service': 'DD Soubre'},
    {'nom': 'DANHO', 'prenom': 'Jocelyn Moise', 'matricule': '3823-ME', 'service': 'DAFP'},
    {'nom': 'DENAN', 'prenom': 'Koffi Crepin', 'matricule': '3824-ME', 'service': 'DD Grand-Bassam'},
    {'nom': 'DIABY', 'prenom': 'Fatoumata Bintou', 'matricule': '3825-ME', 'service': 'DD Dabou'},
    {'nom': 'DIABY', 'prenom': 'Fatoumata Epse KANO', 'matricule': '3826-ME', 'service': 'Cabinet DAF'},
    {'nom': 'DIARRASSOUBA', 'prenom': 'Adama', 'matricule': '3827-ME', 'service': 'DR Divo'},
    {'nom': 'DIARRASSOUBA', 'prenom': 'Moussa', 'matricule': '3828-ME', 'service': 'DR Abidjan'},
    {'nom': 'DOSSO', 'prenom': 'Aboubacar Hamed', 'matricule': '3830-ME', 'service': 'DD Biankouma'},
    {'nom': 'DOUYOU', 'prenom': 'Ines Melissa', 'matricule': '3831-ME', 'service': 'DD Dabou'},
    {'nom': 'DRISSA', 'prenom': 'Anna Mariama', 'matricule': '3832-ME', 'service': 'DDPE'},
    {'nom': 'EFFOUSSOU', 'prenom': 'Suzanne Epse LOBOUE', 'matricule': '3833-ME', 'service': 'DR Agboville'},
    {'nom': 'EHORA', 'prenom': 'Grambe Jean-Marcel', 'matricule': '3834-ME', 'service': 'DR Divo'},
    {'nom': 'EL-HADJ', 'prenom': 'Ouattara Fetigue', 'matricule': '3835-ME', 'service': 'DD Dabou'},
    {'nom': 'ESMEL', 'prenom': 'Edwige Rosa', 'matricule': '3836-ME', 'service': 'DR Agboville'},
    {'nom': 'ESSOH', 'prenom': 'Diby Emmanuel', 'matricule': '3837-ME', 'service': 'DD Agnibilekrou'},
    {'nom': 'GALE', 'prenom': 'Djekou Atteh Antoine', 'matricule': '3838-ME', 'service': 'DR Divo'},
    {'nom': 'GBA', 'prenom': 'Kamingan Tatiana Ange Manuella', 'matricule': '3839-ME', 'service': 'DD Bocanda'},
    {'nom': 'GBOHOU', 'prenom': 'Disseka Diane Epse KONIN', 'matricule': '3840-ME', 'service': 'DPE'},
    {'nom': 'GBANE', 'prenom': 'Bakary', 'matricule': '3841-ME', 'service': 'DD Boundiali'},
    {'nom': 'GLE', 'prenom': 'Ysseuka Sandrine Precilia Epse HORO', 'matricule': '3842-ME', 'service': 'DR Yamoussoukro'},
    {'nom': 'GNAGNE', 'prenom': 'Akpa Jean-Luc', 'matricule': '3843-ME', 'service': 'DD Issia'},
    {'nom': 'GNAHORE', 'prenom': 'Alfred Chantal Valerie Epse LAVRI', 'matricule': '3844-ME', 'service': 'DDPE'},
    {'nom': 'GNAORE', 'prenom': 'Ita Marina', 'matricule': '3845-ME', 'service': 'DD Tiassale'},
    {'nom': 'GNAORE', 'prenom': 'Nina Joelle', 'matricule': '3846-ME', 'service': 'DR San Pedro'},
    {'nom': 'GNOLEBA', 'prenom': 'Daniel Allele', 'matricule': '3847-ME', 'service': 'DD Biolequin'},
    {'nom': 'GNOSSIAN', 'prenom': 'Ange Christelle', 'matricule': '3848-ME', 'service': 'Cabinet'},
    {'nom': 'GOBOU', 'prenom': 'Brudinge', 'matricule': '3849-ME', 'service': 'DR Agboville'},
    {'nom': 'GOGOUA', 'prenom': 'Annick Marcelle', 'matricule': '3850-ME', 'service': 'DR Abidjan'},
    {'nom': 'GOH', 'prenom': 'Ange Christelle Tatiana', 'matricule': '3851-ME', 'service': 'DR Abidjan'},
    {'nom': 'GOREH', 'prenom': 'Mireille Dominique', 'matricule': '3852-ME', 'service': 'DGIR'},
    {'nom': 'GROGUHE', 'prenom': 'Dieudonne Didier', 'matricule': '3853-ME', 'service': 'DAFP'},
    {'nom': 'GROUPESSIE', 'prenom': 'Victor', 'matricule': '3854-ME', 'service': 'DAFP'},
    {'nom': 'GUEHI', 'prenom': 'Toudie Elie Junior Saint-Cyr', 'matricule': '3855-ME', 'service': 'DGIR'},
    {'nom': 'GUEU', 'prenom': 'Pya Hermann', 'matricule': '3856-ME', 'service': 'DR Bouake'},
    {'nom': 'HAIDARA', 'prenom': 'Assata', 'matricule': '3857-ME', 'service': 'DR Agboville'},
    {'nom': 'HOUSSOU', 'prenom': 'Amani Raoul', 'matricule': '3858-ME', 'service': 'DPE'},
    {'nom': 'HUE', 'prenom': 'Lou Bolt Lydie Christine', 'matricule': '3859-ME', 'service': 'DR Tiassale'},
    {'nom': 'IRIE BI', 'prenom': 'Irie Georges Arnaud', 'matricule': '3860-ME', 'service': 'DGIR'},
    {'nom': 'KAMAGATE', 'prenom': 'El Hadj Baba', 'matricule': '3861-ME', 'service': 'DGIR'},
    {'nom': 'KAMAGATE', 'prenom': 'Youssouf Bakir', 'matricule': '3862-ME', 'service': 'DD Zuenoula'},
    {'nom': 'KAMATE', 'prenom': 'Fatoumata', 'matricule': '3863-ME', 'service': 'DD Zuenoula'},
    {'nom': 'KAMENAN', 'prenom': 'Antoine', 'matricule': '3864-ME', 'service': 'DD Tiassale'},
    {'nom': 'KASSI', 'prenom': 'Amon Chantal', 'matricule': '3865-ME', 'service': 'DR Adzope'},
    {'nom': 'KOFFI', 'prenom': 'Adja Gnangoran Estelle', 'matricule': '3867-ME', 'service': 'DRH'},
    {'nom': 'KOFFI', 'prenom': 'Aime Kouadio', 'matricule': '3868-ME', 'service': 'DD Duekoue'},
    {'nom': 'KOFFI', 'prenom': 'Amon Marceline', 'matricule': '3869-ME', 'service': 'DR Aboisso'},
    {'nom': 'KOFFI', 'prenom': 'Aya Elisabeth', 'matricule': '3870-ME', 'service': 'DD Abidjan'},
    {'nom': 'KOFFI', 'prenom': 'Odoukou Marie Florence', 'matricule': '3871-ME', 'service': 'DGIR'},
    {'nom': 'KOKO', 'prenom': 'Aida Judicaelle', 'matricule': '3872-ME', 'service': 'DR Abidjan'},
    {'nom': 'KONAN', 'prenom': 'Adjoua Marinette', 'matricule': '3873-ME', 'service': 'DD Jacqueville'},
    {'nom': 'KONAN', 'prenom': 'Ahouet Nina Chantal', 'matricule': '3874-ME', 'service': 'DDPE'},
    {'nom': 'KOUAME', 'prenom': 'Abou Bah Suzanne Epse', 'matricule': '3875-ME', 'service': 'DR Bouafle'},
    {'nom': 'KONAN', 'prenom': 'Abossoua Anastasie', 'matricule': '3876-ME', 'service': 'DR Abidjan'},
    {'nom': 'KONAN', 'prenom': 'Kenga Bernard', 'matricule': '3877-ME', 'service': 'Cabinet'},
    {'nom': 'KONE', 'prenom': 'Cheick Ibrahim', 'matricule': '3878-ME', 'service': 'DR Divo'},
    {'nom': 'KONE', 'prenom': 'Kenissa Djeneba', 'matricule': '3879-ME', 'service': 'DR Abidjan'},
    {'nom': 'KONE', 'prenom': 'Lancine', 'matricule': '3880-ME', 'service': 'DR Abidjan'},
    {'nom': 'KONE', 'prenom': 'Maimouna', 'matricule': '3881-ME', 'service': 'DD Jacqueville'},
    {'nom': 'KONE', 'prenom': 'Salimata', 'matricule': '3882-ME', 'service': 'DD Lakota'},
    {'nom': 'KONE', 'prenom': 'Sanata', 'matricule': '3883-ME', 'service': 'DR Bondoukou'},
    {'nom': 'KOUADIO', 'prenom': 'Aiffoue Hortense', 'matricule': '3884-ME', 'service': 'DD Beoumi'},
    {'nom': 'KOUADIO', 'prenom': 'Konan Armand', 'matricule': '3886-ME', 'service': 'DD Jacqueville'},
    {'nom': 'KOUADIO', 'prenom': 'Konan Michelle Epse CAMARA', 'matricule': '3887-ME', 'service': 'DD Korhogo'},
    {'nom': 'KOUADIO', 'prenom': 'Loukou Claude', 'matricule': '3888-ME', 'service': 'DD Issia'},
    {'nom': 'KOUADIO', 'prenom': 'Sebastien Ataici', 'matricule': '3889-ME', 'service': 'DAFP'},
    {'nom': 'KOUAKOU', 'prenom': 'Sylvane Epiphanie', 'matricule': '3890-ME', 'service': 'DR Yamoussoukro'},
    {'nom': 'KOUAME', 'prenom': 'Abongoh Leopold', 'matricule': '3891-ME', 'service': 'DAFP'},
    {'nom': 'KOUAME', 'prenom': 'Iya Eugenie', 'matricule': '3892-ME', 'service': 'DD Lakota'},
    {'nom': 'KOUAME', 'prenom': 'Esse Gwindys Desiree Angela', 'matricule': '3893-ME', 'service': 'DD Tiassale'},
    {'nom': 'KOUASSI', 'prenom': 'Yao Etson Alfred', 'matricule': '3894-ME', 'service': 'DD Alepe'},
    {'nom': 'KOUASSI', 'prenom': 'Adjoua Natacha', 'matricule': '3895-ME', 'service': 'DDPE'},
    {'nom': 'KOUASSI', 'prenom': 'Atou Leontine', 'matricule': '3896-ME', 'service': 'DD Grand-Lahou'},
    {'nom': 'KOUASSI', 'prenom': 'Iya Rosalie', 'matricule': '3897-ME', 'service': 'DGIQ'},
    {'nom': 'KOUASSI', 'prenom': 'Eoni Lome Kassidy', 'matricule': '3898-ME', 'service': 'DD Adzope'},
    {'nom': 'KOUASSI', 'prenom': 'Kouame Fabrice Michael', 'matricule': '3899-ME', 'service': 'DR Yamoussoukro'},
    {'nom': 'KOUASSI', 'prenom': 'Yao Serge Vidal', 'matricule': '3900-ME', 'service': 'DD Bangolo'},
    {'nom': 'KPAZAI', 'prenom': 'Jacques Tafarin Douho', 'matricule': '3901-ME', 'service': 'DR Bouafle'},
    {'nom': 'KRA', 'prenom': 'Kossia Estelle', 'matricule': '3902-ME', 'service': 'DAFP'},
    {'nom': 'KRAMO', 'prenom': 'Konan Ahou Beatrice Epse BOHOUSSOU', 'matricule': '3903-ME', 'service': 'DR Divo'},
    {'nom': 'KRAMA', 'prenom': 'Yao Jean Francois', 'matricule': '3904-ME', 'service': 'DR Abengourou'},
    {'nom': 'LAMAH', 'prenom': 'Miriam', 'matricule': '3905-ME', 'service': 'DR Yamoussoukro'},
    {'nom': 'LASSOUMANI', 'prenom': 'Kouakou Joseph', 'matricule': '3906-ME', 'service': 'DAFP'},
    {'nom': 'LELEPO', 'prenom': 'Api Cecile', 'matricule': '3907-ME', 'service': 'DD Lakota'},
    {'nom': 'LOLE', 'prenom': 'Narcisse', 'matricule': '3908-ME', 'service': 'DD Tieboissou'},
    {'nom': 'MABA', 'prenom': 'Zinan Marie Rosine', 'matricule': '3909-ME', 'service': 'DD Grand-Lahou'},
    {'nom': 'MALANON', 'prenom': 'Kouakou Kpagni Jean-Pierre', 'matricule': '3910-ME', 'service': 'DD Tiebissou'},
    {'nom': 'MAMBO', 'prenom': 'Axelle Audrey', 'matricule': '3911-ME', 'service': 'DD Alepe'},
    {'nom': 'MAMDET', 'prenom': 'Paule Anicette Lyas', 'matricule': '3912-ME', 'service': 'DGIQ'},
    {'nom': 'MBRA', 'prenom': 'Aya Marie Gisele', 'matricule': '3913-ME', 'service': 'DD Adiake'},
    {'nom': 'MEL', 'prenom': "N'Guessan Herve Jacques", 'matricule': '3914-ME', 'service': 'DD Dabou'},
    {'nom': 'MEL', 'prenom': 'Meless Aurelie Jeannette', 'matricule': '3915-ME', 'service': 'DR Man'},
    {'nom': 'MOBIO', 'prenom': 'Nita Natacha', 'matricule': '3916-ME', 'service': 'DD Jacqueville'},
    {'nom': 'MONSAN', 'prenom': 'Jpo Sylvie', 'matricule': '3917-ME', 'service': 'DR Adzope'},
    {'nom': 'NEKALO', 'prenom': 'Goliva Edith Arnaud', 'matricule': '3918-ME', 'service': 'DR Adzope'},
    {'nom': "N'GUESSAN", 'prenom': 'Koffi Elisee', 'matricule': '3919-ME', 'service': 'DD Alepe'},
    {'nom': 'NIENFA', 'prenom': 'Nagore Saint Claire', 'matricule': '3920-ME', 'service': 'DD Jacqueville'},
    {'nom': 'ODETTE', 'prenom': 'Ouabio', 'matricule': '3921-ME', 'service': 'DD Jacqueville'},
    {'nom': 'OKEI', 'prenom': 'Okei Junior Pascal', 'matricule': '3922-ME', 'service': 'DD Adiake'},
    {'nom': 'OKPO', 'prenom': 'Armelle Aurelie', 'matricule': '3923-ME', 'service': 'DD Jacqueville'},
    {'nom': 'OSSORON', 'prenom': 'Diane Marina Epse NHOUMI', 'matricule': '3924-ME', 'service': 'DD Tiassale'},
    {'nom': 'OUADA', 'prenom': 'Alloue Melanie Epse SAHIE', 'matricule': '3925-ME', 'service': 'DRH'},
    {'nom': 'OUATTARA', 'prenom': 'Adjoumanin Koffi Amed', 'matricule': '3926-ME', 'service': 'DR Agboville'},
    {'nom': 'OUATTARA', 'prenom': 'Ahmed Al Hassan Abissa', 'matricule': '3927-ME', 'service': 'DDPE'},
    {'nom': 'OUATTARA', 'prenom': 'Dogrounani', 'matricule': '3928-ME', 'service': 'DR Abengourou'},
    {'nom': 'OUATTARA', 'prenom': 'Dogrounani', 'matricule': '3929-ME', 'service': 'DD Dabou'},
    {'nom': 'OUATTARA', 'prenom': 'Koffi Ali', 'matricule': '3930-ME', 'service': 'DD Issia'},
    {'nom': 'OUATTARA', 'prenom': 'Kouabenan Siedou', 'matricule': '3931-ME', 'service': 'DR Bondoukou'},
    {'nom': 'OUATTARA', 'prenom': 'Sekou', 'matricule': '3932-ME', 'service': 'DR Abengourou'},
    {'nom': 'QUEHE', 'prenom': 'Tcheblea Eunice', 'matricule': '3933-ME', 'service': 'DD Lakota'},
    {'nom': 'PAGNE', 'prenom': 'Mireille Angeline Mondesir', 'matricule': '3934-ME', 'service': 'DGIR'},
    {'nom': 'PANGNY', 'prenom': 'Ange Camille', 'matricule': '3935-ME', 'service': 'DD Abidjan'},
    {'nom': 'SAHI', 'prenom': 'Finde Sinn', 'matricule': '3936-ME', 'service': 'DR Abengourou'},
    {'nom': 'SANGARE', 'prenom': 'Awa', 'matricule': '3937-ME', 'service': 'DD Jacqueville'},
    {'nom': 'SIGNO', 'prenom': 'Abenan Sonia', 'matricule': '3940-ME', 'service': 'DD Tanda'},
    {'nom': 'SOULEYMANE', 'prenom': 'Harouna', 'matricule': '3941-ME', 'service': 'DR Yamoussoukro'},
    {'nom': 'TANOH', 'prenom': 'Sopy Henriette', 'matricule': '3942-ME', 'service': 'DR Divo'},
    {'nom': 'TIBE LOU', 'prenom': 'Clai Berthie Epse BOLY', 'matricule': '3943-ME', 'service': 'DRH'},
    {'nom': 'TOHOUA', 'prenom': 'Gadyon Video', 'matricule': '3944-ME', 'service': 'DD Babakala'},
    {'nom': 'TRAORE', 'prenom': 'Moussa Pie', 'matricule': '3945-ME', 'service': 'DPE'},
    {'nom': 'TRAORE', 'prenom': 'Mariam', 'matricule': '3946-ME', 'service': 'DD Tanda'},
    {'nom': 'WANTAN', 'prenom': 'Afoua Folie Brigitte', 'matricule': '3947-ME', 'service': 'DD Tanda'},
    {'nom': 'YAO', 'prenom': 'Afoue Valentine', 'matricule': '3949-ME', 'service': 'DD Jacqueville'},
    {'nom': 'YAO', 'prenom': 'Klama Anne-Marie', 'matricule': '3950-ME', 'service': 'DR Bondoukou'},
    {'nom': 'YAO', 'prenom': 'Koffi Fomaric', 'matricule': '3951-ME', 'service': 'DD Jacqueville'},
    {'nom': 'YAO', 'prenom': 'Kouame dit Souleymane', 'matricule': '3952-ME', 'service': 'DAFP'},
    {'nom': 'YAO', 'prenom': 'Kouane Franck', 'matricule': '3953-ME', 'service': 'Cabinet'},
    {'nom': 'YAO', 'prenom': 'Ouattara Aboubacar', 'matricule': '3954-ME', 'service': 'DD Grand-Bassam'},
    {'nom': 'YAPO', 'prenom': 'Ademon Sandrine Marie-France', 'matricule': '3955-ME', 'service': 'DR Divo'},
    {'nom': 'YAPO', 'prenom': 'Konan Junior Elvis', 'matricule': '3957-ME', 'service': 'DR Abidjan'},
    {'nom': 'YAPOCA', 'prenom': 'Laurette Daniela', 'matricule': '3958-ME', 'service': 'DRH'},
    {'nom': 'YAVO', 'prenom': "M'Boya Sainte Flora d'Angelus", 'matricule': '3959-ME', 'service': 'DD Adiake'},
    {'nom': 'ZOUBI', 'prenom': 'Diagone Paul Ambroise', 'matricule': '3960-ME', 'service': 'DD Alepe'},
    {'nom': 'BEUGRE', 'prenom': 'Bassy Jean-Michel', 'matricule': '6029-ME', 'service': 'DAFP'},
    {'nom': 'YOUAN', 'prenom': 'Lou Goure Hortense', 'matricule': '6030-ME', 'service': 'DR Agboville'},
    {'nom': 'COULIBALY', 'prenom': 'Salimata Valerie', 'matricule': '6031-ME', 'service': 'DR Agboville'},
    {'nom': 'GNACHOUE', 'prenom': 'Ade Christine Zita', 'matricule': '6032-ME', 'service': 'DD Bongouanou'},
    {'nom': 'RYAN', 'prenom': 'Gneka', 'matricule': '6033-ME', 'service': 'DD Grand-Lahou'},
    {'nom': 'TRA', 'prenom': 'Lou Gounan Eleonor Epse OUATTARA', 'matricule': '6034-ME', 'service': 'DD Danane'},
    {'nom': 'SAMAKE', 'prenom': 'Mariam', 'matricule': '6035-ME', 'service': 'DRH'},
    {'nom': "N'GUESSAN", 'prenom': 'Antoin Flore Liliane', 'matricule': '6036-ME', 'service': 'DDPE'},
    {'nom': 'BONI', 'prenom': 'Chiakin Melaine Patricia', 'matricule': '6037-ME', 'service': 'DR Adzope'},
    {'nom': 'ABOUA', 'prenom': 'Prisca Melaine', 'matricule': '6038-ME', 'service': 'DD Alepe'},
    {'nom': 'KOFFI', 'prenom': 'Konan Sylvain', 'matricule': '6039-ME', 'service': 'DDPE'},
    {'nom': 'BONZOU', 'prenom': 'Kouame Amakon Silva', 'matricule': '6040-ME', 'service': 'DAF'},
]

def generate_username(prenom, nom, existing_usernames):
    first = prenom.split()[0].lower().replace("'", "").replace("-", "").replace(".", "")
    last = nom.split()[0].lower().replace("'", "").replace("-", "").replace(".", "")
    username = f"{first}.{last}"
    if username in existing_usernames:
        i = 2
        while f"{username}{i}" in existing_usernames:
            i += 1
        username = f"{username}{i}"
    return username

def generate_password(nom, index):
    clean_nom = nom.split()[0].replace("'", "").replace("-", "")
    prefix = clean_nom[:4].capitalize()
    return f"Azin@{prefix}{index:02d}"

print("=" * 70)
print("CREATION DES 173 COMPTES AGENTS AZING IVOIR (AGENTS DE BUREAU)")
print("=" * 70)

existing_usernames = set(User.objects.values_list('username', flat=True))
created = 0
skipped = 0
results = []

for idx, agent in enumerate(agents, 1):
    username = generate_username(agent['prenom'], agent['nom'], existing_usernames)
    password = generate_password(agent['nom'], idx)

    if User.objects.filter(username=username).exists():
        print(f"  [{idx:3d}] DEJA EXISTANT : {username}")
        skipped += 1
        continue

    user = User.objects.create_user(
        username=username,
        password=password,
        first_name=agent['prenom'],
        last_name=agent['nom'],
        email=f"{username}@azing-ivoir.ci",
        is_staff=False,
        is_superuser=False
    )

    # Mettre a jour l'employe existant ou en creer un nouveau
    emp_qs = Employee.objects.filter(matricule=agent['matricule'])
    if emp_qs.exists():
        emp = emp_qs.first()
        emp.user = user
        emp.save()
    else:
        Employee.objects.create(
            user=user,
            first_name=agent['prenom'],
            last_name=agent['nom'],
            email=f"{username}@azing-ivoir.ci",
            matricule=agent['matricule'],
            department=azing,
            direction=agent['service'],
            position='Agent de bureau',
            hire_date='2025-12-01',
            status='active'
        )

    PasswordRecord.objects.update_or_create(
        user=user,
        defaults={
            'password_plain': password,
            'role': 'employee',
        }
    )

    existing_usernames.add(username)
    created += 1
    results.append({
        'num': idx,
        'nom_prenom': f"{agent['nom']} {agent['prenom']}",
        'username': username,
        'password': password,
        'poste': 'Agent de bureau',
        'service': agent['service'],
    })
    print(f"  [{idx:3d}] {agent['nom']} {agent['prenom']}")
    print(f"       Username: {username} | Password: {password} | Service: {agent['service']}")

print(f"\n{'=' * 70}")
print(f"RESULTAT : {created} comptes crees, {skipped} ignores")
print(f"{'=' * 70}")

# === GENERATION DU FICHIER EXCEL ===
if results:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Identifiants AZING IVOIR"

    # Titre
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "AZING IVOIR Sarl - TABLEAU DES IDENTIFIANTS ET MOTS DE PASSE"
    t.font = Font(bold=True, size=14, color="000000")
    t.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value = "Prestation de service - Agents contractuels - Decembre 2025"
    s.font = Font(italic=True, size=10, color="666666")
    s.alignment = Alignment(horizontal="center")

    # En-tetes
    headers = ["N°", "NOM ET PRENOMS", "IDENTIFIANT", "MOT DE PASSE", "POSTE", "SERVICE"]
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Donnees
    for i, r in enumerate(results):
        row = i + 5
        vals = [r['num'], r['nom_prenom'], r['username'], r['password'], r['poste'], r['service']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 25

    filepath = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Identifiants_AZING_IVOIR.xlsx')
    # Fallback: save next to the script
    filepath2 = os.path.join(os.path.dirname(__file__), '..', 'Identifiants_AZING_IVOIR.xlsx')
    try:
        wb.save(filepath)
        print(f"\nFichier Excel cree : {filepath}")
    except Exception:
        wb.save(filepath2)
        print(f"\nFichier Excel cree : {filepath2}")
