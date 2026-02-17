"""
Script pour exporter la table des mots de passe en fichier Excel.
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'empmanager.settings')
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib.auth.models import User
from api.models import PasswordRecord, Employee

wb = Workbook()
ws = wb.active
ws.title = "Identifiants"

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(horizontal='left', vertical='center')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Titre
ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = 'TABLEAU DES IDENTIFIANTS ET MOTS DE PASSE'
title_cell.font = Font(name='Arial', bold=True, size=14, color='2E7D32')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# Ligne vide
ws.row_dimensions[2].height = 10

# En-tetes
headers = ['N\u00b0', 'NOM ET PRENOMS', 'IDENTIFIANT', 'MOT DE PASSE', 'ROLE', 'ENTREPRISE', 'DIRECTION']
col_widths = [6, 35, 25, 20, 18, 22, 30]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col_idx)].width = width

ws.row_dimensions[3].height = 25

# Donnees
records = PasswordRecord.objects.select_related('user').all().order_by('role', 'user__last_name', 'user__first_name')

role_labels = {
    'admin': 'Administrateur',
    'manager': 'Manager',
    'employee': 'Employe',
}

# Couleurs par role
role_fills = {
    'admin': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
    'manager': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
    'employee': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
}

row_num = 4
for idx, record in enumerate(records, 1):
    user = record.user
    full_name = user.get_full_name() or user.username

    # Chercher l'entreprise et la direction
    dept_name = '-'
    direction = '-'
    try:
        emp = user.employee_profile
        if emp.department:
            dept_name = emp.department.name
        if emp.direction:
            direction = emp.direction
    except Employee.DoesNotExist:
        pass

    row_data = [
        idx,
        full_name.upper(),
        record.user.username,
        record.password_plain,
        role_labels.get(record.role, record.role),
        dept_name,
        direction,
    ]

    row_fill = role_fills.get(record.role, role_fills['employee'])

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border
        cell.fill = row_fill
        if col_idx == 1:
            cell.alignment = center_align
        elif col_idx == 5:
            cell.alignment = center_align
        else:
            cell.alignment = cell_align

    ws.row_dimensions[row_num].height = 20
    row_num += 1

# Ligne de total
ws.merge_cells(f'A{row_num}:D{row_num}')
total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL : {len(records)} utilisateurs')
total_cell.font = Font(name='Arial', bold=True, size=11)
total_cell.alignment = Alignment(horizontal='center', vertical='center')
total_cell.border = thin_border
for col in range(2, 8):
    ws.cell(row=row_num, column=col).border = thin_border

# Sauvegarder
base_dir = os.path.dirname(os.path.abspath(os.getcwd()))
output_path = os.path.join(base_dir, 'Identifiants_Mots_de_Passe.xlsx')
wb.save(output_path)
print(f"Fichier Excel cree : {output_path}")
print(f"Total : {len(records)} utilisateurs exportes")
