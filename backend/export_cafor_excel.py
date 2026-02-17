"""
Script pour exporter les identifiants des agents CAFOR en fichier Excel.
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'empmanager.settings')
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from api.models import PasswordRecord, Employee

wb = Workbook()
ws = wb.active
ws.title = "Agents CAFOR"

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='8D6E63', end_color='8D6E63', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(horizontal='left', vertical='center')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ws.merge_cells('A1:G1')
title_cell = ws['A1']
title_cell.value = 'CAFOR - TABLEAU DES IDENTIFIANTS ET MOTS DE PASSE'
title_cell.font = Font(name='Arial', bold=True, size=14, color='5D4037')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:G2')
sub_cell = ws['A2']
sub_cell.value = 'Centre Autonome de Formation de Recyclage et de Prestations - Chauffeurs - Decembre 2025'
sub_cell.font = Font(name='Arial', italic=True, size=10, color='666666')
sub_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 20

ws.row_dimensions[3].height = 8

headers = ['N\u00b0', 'NOM ET PRENOMS', 'IDENTIFIANT', 'MOT DE PASSE', 'POSTE', 'SERVICE']
col_widths = [6, 38, 25, 20, 14, 25]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col_idx)].width = width

ws.row_dimensions[4].height = 25

employees = Employee.objects.filter(
    department__name='CAFOR'
).select_related('user', 'department').order_by('last_name', 'first_name')

row_num = 5
count = 0

for idx, emp in enumerate(employees, 1):
    if not emp.user:
        continue

    password = '(inconnu)'
    try:
        record = PasswordRecord.objects.get(user=emp.user)
        password = record.password_plain
    except PasswordRecord.DoesNotExist:
        pass

    full_name = f"{emp.last_name} {emp.first_name}"

    row_data = [
        idx,
        full_name.upper(),
        emp.user.username,
        password,
        emp.position,
        emp.direction or '-',
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border
        if col_idx in (1, 5):
            cell.alignment = center_align
        else:
            cell.alignment = cell_align

    ws.row_dimensions[row_num].height = 20
    row_num += 1
    count += 1

ws.merge_cells(f'A{row_num}:D{row_num}')
total_cell = ws.cell(row=row_num, column=1, value=f'TOTAL : {count} agents')
total_cell.font = Font(name='Arial', bold=True, size=11)
total_cell.alignment = Alignment(horizontal='center', vertical='center')
total_cell.border = thin_border
for col in range(2, 7):
    ws.cell(row=row_num, column=col).border = thin_border

base_dir = os.path.dirname(os.path.abspath(os.getcwd()))
output_path = os.path.join(base_dir, 'Identifiants_CAFOR.xlsx')
wb.save(output_path)
print(f"Fichier Excel cree : {output_path}")
print(f"Total : {count} agents CAFOR exportes")
