"""
Script pour créer le fichier Excel des agents AZING IVOIR
et les intégrer dans la base de données Django
"""
import os
import sys
import django

# Données des 173 agents extraites de la liste de présence AZING IVOIR - Décembre 2025
AGENTS = [
    (1, "3791-ME", "ABO", "Brou Juliette", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (2, "3792-ME", "ABOH", "Tano François", "Agent de bureau", "Direction Départementale (DD) de Bongouanou"),
    (3, "3793-ME", "ACAFOU", "Hyacinthe", "Agent de bureau", "Direction Régionale (DR) d'Alépé"),
    (4, "3794-ME", "ACHI", "Chantal Georgette", "Agent de bureau", "Direction Départementale (DD) d'Agboville"),
    (5, "3795-ME", "ACHO", "Apo Marie Chantal", "Agent de bureau", "Direction Départementale (DD) de Tanda"),
    (6, "3796-ME", "ADOUKO", "Koussi Firmin", "Agent de bureau", "Direction Départementale (DD) d'Abidjan"),
    (7, "3797-ME", "ADJOBI", "Narcisse", "Agent de bureau", "Direction Départementale (DD) de Soubré"),
    (8, "3798-ME", "AKE", "N'Cho Johness Eric", "Agent de bureau", "Direction Départementale (DD) de Zuénoula"),
    (9, "3800-ME", "AMANI", "Konan", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (10, "3801-ME", "AMARA", "Ouattara", "Agent de bureau", "Direction Départementale (DD) de Grand-Bassam"),
    (11, "3802-ME", "ASSOUMO", "J. Eric", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (12, "3803-ME", "ATSE", "Yapo Henri Didier", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (13, "3804-ME", "AVEREBY", "Katie Phyrace", "Agent de bureau", "Direction Départementale (DD) d'Oumé"),
    (14, "3805-ME", "BAMBA", "Fessimata", "Agent de bureau", "Direction Régionale (DR) de Man"),
    (15, "3807-ME", "BEDA", "Nadine Micheline", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (16, "3808-ME", "BEHO", "Djerou Agnès Epse BAH", "Agent de bureau", "Direction Départementale (DD) de Sassandra"),
    (17, "3809-ME", "BLEURIEZ", "Guiho Erwige Kevine Epse KOUASSI", "Agent de bureau", "Inspection Générale"),
    (18, "3810-ME", "BLI", "Fatou Patricia", "Agent de bureau", "Direction Régionale (DR) de Korhogo"),
    (19, "3811-ME", "BOKA", "Konan Anastasie", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (20, "3812-ME", "BONI", "Assi Gildas Thibaut", "Agent de bureau", "Direction Départementale (DD) d'Issia"),
    (21, "3813-ME", "BOUCHOU", "Placide Allali Didier Gilbert", "Agent de bureau", "Direction Départementale (DD) de Tiassalé"),
    (22, "3814-ME", "BOUSSOU", "Armelie Rosine", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (23, "3815-ME", "BROU", "Alla Koko Jeannette", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (24, "3816-ME", "BROU", "N'Guessan Noelle", "Agent de bureau", "Direction Départementale (DD) de Touba"),
    (25, "3817-ME", "BROU", "Obehi Chaye Melaine Annick", "Agent de bureau", "Direction des Ressources Humaines (DRH)"),
    (26, "3818-ME", "CAMARA", "Abdoulaye", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (27, "3819-ME", "CHIMENE", "Nanou Eichem", "Agent de bureau", "Direction Régionale (DR) d'Aboisso"),
    (28, "3820-ME", "COULIBALY", "Yopie Awa", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (29, "3821-ME", "DAGNOKO", "Aminata", "Agent de bureau", "Direction de la Planification et de l'Évaluation (DPE)"),
    (30, "3822-ME", "DAGNOKO", "Mariam", "Agent de bureau", "Direction Départementale (DD) de Soubré"),
    (31, "3823-ME", "DANHO", "Jocelyn Moïse", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (32, "3824-ME", "DENAN", "Koffi Crépin", "Agent de bureau", "Direction Départementale (DD) de Grand-Bassam"),
    (33, "3825-ME", "DIABY", "Fatoumata Bintou", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (34, "3826-ME", "DIABY", "Fatoumata Epse KANO", "Agent de bureau", "Cabinet Direction des Affaires Financières (DAF)"),
    (35, "3827-ME", "DIARRASSOUBA", "Adama", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (36, "3828-ME", "DIARRASSOUBA", "Moussa", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (37, "3830-ME", "DOSSO", "Aboubacar Hamed", "Agent de bureau", "Direction Départementale (DD) de Biankouma"),
    (38, "3831-ME", "DOUYOU", "Ines Melissa", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (39, "3832-ME", "DRISSA", "Anna Mariama", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (40, "3833-ME", "EFFOUSSOU", "Suzanne Epse LOBOUE", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (41, "3834-ME", "EHORA", "Grambe Jean-Marcel", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (42, "3835-ME", "EL-HADJ", "Ouattara Fetigue", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (43, "3836-ME", "ESMEL", "Edwige Rosa", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (44, "3837-ME", "ESSOH", "Diby Emmanuel", "Agent de bureau", "Direction Départementale (DD) d'Agnibilékrou"),
    (45, "3838-ME", "GALE", "Djékou Atteh Antoine", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (46, "3839-ME", "GBA", "Kamingan Tatiana Ange Manuella", "Agent de bureau", "Direction Départementale (DD) de Bocanda"),
    (47, "3840-ME", "GBOHOU", "Disseka Diane Epse KONIN", "Agent de bureau", "Direction de la Planification et de l'Évaluation (DPE)"),
    (48, "3841-ME", "GBANE", "Bakary", "Agent de bureau", "Direction Départementale (DD) de Boundiali"),
    (49, "3842-ME", "GLE", "Ysseuka Sandrine Precilia Epse HORO", "Agent de bureau", "Direction Régionale (DR) de Yamoussoukro"),
    (50, "3843-ME", "GNAGNE", "Akpa Jean-Luc", "Agent de bureau", "Direction Départementale (DD) d'Issia"),
    (51, "3844-ME", "GNAHORE", "Alfred Chantal Valerie Epse LAVRI", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (52, "3845-ME", "GNAORE", "Ita Marina", "Agent de bureau", "Direction Départementale (DD) de Tiassalé"),
    (53, "3846-ME", "GNAORE", "Nina Joëlle", "Agent de bureau", "Direction Régionale (DR) de San Pedro"),
    (54, "3847-ME", "GNOLEBA", "Daniel Allélé", "Agent de bureau", "Direction Départementale (DD) de Bloléquin"),
    (55, "3848-ME", "GNOSSIAN", "Ange Christelle", "Agent de bureau", "Cabinet"),
    (56, "3849-ME", "GOBOU", "Brudinge", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (57, "3850-ME", "GOGOUA", "Annick Marcelle", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (58, "3851-ME", "GOH", "Ange Christelle Tatiana", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (59, "3852-ME", "GOREH", "Mireille Dominique", "Agent de bureau", "Direction Générale des Infrastructures Routières (DGIR)"),
    (60, "3853-ME", "GROGUHE", "Dieudonné Didier", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (61, "3854-ME", "GROUPESSIE", "Victor", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (62, "3855-ME", "GUEHI", "Toudie Elie Junior Saint-Cyr", "Agent de bureau", "Direction Générale des Infrastructures Routières (DGIR)"),
    (63, "3856-ME", "GUEU", "Pya Hermann", "Agent de bureau", "Direction Régionale (DR) de Bouaké"),
    (64, "3857-ME", "HAIDARA", "Assata", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (65, "3858-ME", "HOUSSOU", "Amani Raoul", "Agent de bureau", "Direction de la Planification et de l'Évaluation (DPE)"),
    (66, "3859-ME", "HUE", "Lou Bolt Lydie Christine", "Agent de bureau", "Direction Régionale (DR) de Tiassalé"),
    (67, "3860-ME", "IRIE BI", "Irie Georges Arnaud", "Agent de bureau", "Direction Générale des Infrastructures Routières (DGIR)"),
    (68, "3861-ME", "KAMAGATE", "El Hadj Baba", "Agent de bureau", "Direction Générale des Infrastructures Routières Routières (DGIR)"),
    (69, "3862-ME", "KAMAGATE", "Youssouf Bakir", "Agent de bureau", "Direction Départementale (DD) de Zuénoula"),
    (70, "3863-ME", "KAMATE", "Fatoumata", "Agent de bureau", "Direction Départementale (DD) de Zuénoula"),
    (71, "3864-ME", "KAMENAN", "Antoine", "Agent de bureau", "Direction Départementale (DD) de Tiassalé"),
    (72, "3865-ME", "KASSI", "Amon Chantal", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (73, "3867-ME", "KOFFI", "Adja Gnangoran Estelle", "Agent de bureau", "Direction des Ressources Humaines (DRH)"),
    (74, "3868-ME", "KOFFI", "Aimé Kouadio", "Agent de bureau", "Direction Départementale (DD) de Duékoué"),
    (75, "3869-ME", "KOFFI", "Amon Marceline", "Agent de bureau", "Direction Régionale (DR) d'Aboisso"),
    (76, "3870-ME", "KOFFI", "Aya Elisabeth", "Agent de bureau", "Direction Départementale (DD) d'Abidjan"),
    (77, "3871-ME", "KOFFI", "Odoukou Marie Florence", "Agent de bureau", "Direction Générale des Infrastructures Routières (DGIR)"),
    (78, "3872-ME", "KOKO", "Aida Judicaëlle", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (79, "3873-ME", "KONAN", "Adjoua Marinette", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (80, "3874-ME", "KONAN", "Ahouet Nina Chantal", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (81, "3875-ME", "KOUAME", "Abou Bah Suzanne Epse", "Agent de bureau", "Direction Régionale (DR) de Bouaflé"),
    (82, "3876-ME", "KONAN", "Abossoua Anastasie", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (83, "3877-ME", "KONAN", "Kenga Bernard", "Agent de bureau", "Cabinet"),
    (84, "3878-ME", "KONE", "Cheick Ibrahim", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (85, "3879-ME", "KONE", "Kenissa Djeneba", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (86, "3880-ME", "KONE", "Lanciné", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (87, "3881-ME", "KONE", "Maimouna", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (88, "3882-ME", "KONE", "Salimata", "Agent de bureau", "Direction Départementale (DD) de Lakota"),
    (89, "3883-ME", "KONE", "Sanata", "Agent de bureau", "Direction Régionale (DR) de Bondoukou"),
    (90, "3884-ME", "KOUADIO", "Aiffoue Hortense", "Agent de bureau", "Direction Départementale (DD) de Béoumi"),
    (91, "3886-ME", "KOUADIO", "Konan Armand", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (92, "3887-ME", "KOUADIO", "Konan Michelle Epse CAMARA", "Agent de bureau", "Direction Départementale (DD) de Korhogo"),
    (93, "3888-ME", "KOUADIO", "Loukou Claude", "Agent de bureau", "Direction Départementale (DD) d'Issia"),
    (94, "3889-ME", "KOUADIO", "Sébastien Ataïci", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (95, "3890-ME", "KOUAKOU", "Sylvane Epiphanie", "Agent de bureau", "Direction Régionale (DR) de Yamoussoukro"),
    (96, "3891-ME", "KOUAME", "Abongoh Léopold", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (97, "3892-ME", "KOUAME", "Iya Eugénie", "Agent de bureau", "Direction Départementale (DD) de Lakota"),
    (98, "3893-ME", "KOUAME", "Esse Gwindys Désirée Angela", "Agent de bureau", "Direction Départementale (DD) de Tiassalé"),
    (99, "3894-ME", "KOUASSI", "Yao Etson Alfred", "Agent de bureau", "Direction Départementale (DD) d'Alépé"),
    (100, "3895-ME", "KOUASSI", "Adjoua Natacha", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (101, "3896-ME", "KOUASSI", "Atou Léontine", "Agent de bureau", "Direction Départementale (DD) de Grand-Lahou"),
    (102, "3897-ME", "KOUASSI", "Iya Rosalie", "Agent de bureau", "DGIQ"),
    (103, "3898-ME", "KOUASSI", "Eoni Lome Kassidy", "Agent de bureau", "Direction Départementale (DD) d'Adzopé"),
    (104, "3899-ME", "KOUASSI", "Kouamé Fabrice Michael", "Agent de bureau", "Direction Régionale (DR) de Yamoussoukro"),
    (105, "3900-ME", "KOUASSI", "Yao Serge Vidal", "Agent de bureau", "Direction Départementale (DD) de Bangolo"),
    (106, "3901-ME", "KPAZAI", "Jacques Tafarin Douho", "Agent de bureau", "Direction Départementale (DR) de Bouaflé"),
    (107, "3902-ME", "KRA", "Kossia Estelle", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (108, "3903-ME", "KRAMO", "Konan Ahou Béatrice Epse BOHOUSSOU", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (109, "3904-ME", "KRAMA", "Yao Jean François", "Agent de bureau", "Direction Régionale (DR) d'Abengourou"),
    (110, "3905-ME", "LAMAH", "Miriam", "Agent de bureau", "Direction Régionale (DR) de Yamoussoukro"),
    (111, "3906-ME", "LASSOUMANI", "Kouakou Joseph", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (112, "3907-ME", "LELEPO", "Api Cécile", "Agent de bureau", "Direction Départementale (DD) de Lakota"),
    (113, "3908-ME", "LOLE", "Narcisse", "Agent de bureau", "Direction Départementale (DD) de Tiéboissou"),
    (114, "3909-ME", "MABA", "Zinan Marie Rosine", "Agent de bureau", "Direction Départementale (DD) de Grand-Lahou"),
    (115, "3910-ME", "MALANON", "Kouakou Kpagni Jean-Pierre", "Agent de bureau", "Direction Départementale (DD) de Tiébissou"),
    (116, "3911-ME", "MAMBO", "Axelle Audrey", "Agent de bureau", "Direction Départementale (DD) d'Alépé"),
    (117, "3912-ME", "MAMDET", "Paule Anicette Lyas", "Agent de bureau", "DGIQ"),
    (118, "3913-ME", "MBRA", "Aya Marie Gisele", "Agent de bureau", "Direction Départementale (DD) d'Adiaké"),
    (119, "3914-ME", "MEL", "N'Guessan Hervé Jacques", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (120, "3915-ME", "MEL", "Meless Aurélie Jeannette", "Agent de bureau", "Direction Régionale (DR) de Man"),
    (121, "3916-ME", "MOBIO", "Nita Natacha", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (122, "3917-ME", "MONSAN", "Jpo Sylvie", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (123, "3918-ME", "NEKALO", "Goliva Edith Arnaud", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (124, "3919-ME", "N'GUESSAN", "Koffi Elisée", "Agent de bureau", "Direction Départementale (DD) d'Alépé"),
    (125, "3920-ME", "NIENFA", "Nagore Saint Claire", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (126, "3921-ME", "ODETTE", "Ouabio", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (127, "3922-ME", "OKEI", "Okei Junior Pascal", "Agent de bureau", "Direction Départementale (DD) d'Adiaké"),
    (128, "3923-ME", "OKPO", "Armelle Aurelie", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (129, "3924-ME", "OSSORON", "Diane Marina Epse N'HOUMI", "Agent de bureau", "Direction Départementale (DD) de Tiassalé"),
    (130, "3925-ME", "OUADA", "Alloue Melanie Epse SAHIE", "Agent de bureau", "Direction des Ressources Humaines (DRH)"),
    (131, "3926-ME", "OUATTARA", "Adjoumanin Koffi Amed", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (132, "3927-ME", "OUATTARA", "Ahmed Al Hassan Abissa", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (133, "3928-ME", "OUATTARA", "Dogrounani", "Agent de bureau", "Direction Régionale (DR) d'Abengourou"),
    (134, "3929-ME", "OUATTARA", "Dogrounani", "Agent de bureau", "Direction Départementale (DD) de Dabou"),
    (135, "3930-ME", "OUATTARA", "Koffi Ali", "Agent de bureau", "Direction Départementale (DD) d'Issia"),
    (136, "3931-ME", "OUATTARA", "Kouabenan Siedou", "Agent de bureau", "Direction Régionale (DR) de Bondoukou"),
    (137, "3932-ME", "OUATTARA", "Sekou", "Agent de bureau", "Direction Régionale (DR) d'Abengourou"),
    (138, "3933-ME", "QUEHE", "Tcheblea Eunice", "Agent de bureau", "Direction départementale (DD) de Lakota"),
    (139, "3934-ME", "PAGNE", "Mireille Angéline Mondésir", "Agent de bureau", "Direction Générale des Infrastructures Routières (DGIR)"),
    (140, "3935-ME", "PANGNY", "Ange Camille", "Agent de bureau", "Direction Départementale (DD) d'Abidjan"),
    (141, "3936-ME", "SAHI", "Finde Sinn", "Agent de bureau", "Direction Régionale (DR) d'Abengourou"),
    (142, "3937-ME", "SANGARE", "Awa", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (143, "3940-ME", "SIGNO", "Abenan Sonia", "Agent de bureau", "Direction Départementale (DD) de Tanda"),
    (144, "3941-ME", "SOULEYMANE", "Harouna", "Agent de bureau", "Direction Régionale (DR) de Yamoussoukro"),
    (145, "3942-ME", "TANOH", "Sopy Henriette", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (146, "3943-ME", "TIBE LOU", "Clai Berthie Epse BOLY", "Agent de bureau", "Direction des Ressources Humaines (DRH)"),
    (147, "3944-ME", "TOHOUA", "Gadyon Video", "Agent de bureau", "Direction Départementale (DD) de Babakala"),
    (148, "3945-ME", "TRAORE", "Moussa Pie", "Agent de bureau", "Direction de la Planification et de l'Évaluation (DPE)"),
    (149, "3946-ME", "TRAORE", "Mariam", "Agent de bureau", "Direction Départementale (DD) de Tanda"),
    (150, "3947-ME", "WANTAN", "Afoua Folie Brigitte", "Agent de bureau", "Direction Départementale (DD) de Tanda"),
    (151, "3949-ME", "YAO", "Afoue Valentine", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (152, "3950-ME", "YAO", "Klama Anne-Marie", "Agent de bureau", "Direction Régionale (DR) de Bondoukou"),
    (153, "3951-ME", "YAO", "Koffi Fomaric", "Agent de bureau", "Direction Départementale (DD) de Jacqueville"),
    (154, "3952-ME", "YAO", "Kouamé dit Souleymane", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (155, "3953-ME", "YAO", "Kouane Franck", "Agent de bureau", "Cabinet"),
    (156, "3954-ME", "YAO", "Ouattara Aboubacar", "Agent de bureau", "Direction Départementale (DD) de Grand-Bassam"),
    (157, "3955-ME", "YAPO", "Ademon Sandrine Marie-France", "Agent de bureau", "Direction Régionale (DR) de Divo"),
    (158, "3957-ME", "YAPO", "Konan Junior Elvis", "Agent de bureau", "Direction Régionale (DR) d'Abidjan"),
    (159, "3958-ME", "YAPOCA", "Laurette Daniela", "Agent de bureau", "Direction des Ressources Humaines (DRH)"),
    (160, "3959-ME", "YAVO", "M'Boya Sainte Flora d'Angélus", "Agent de bureau", "Direction Départementale (DD) d'Adiaké"),
    (161, "3960-ME", "ZOUBI", "Diagone Paul Ambroise", "Agent de bureau", "Direction départementale (DD) d'Alépé"),
    (162, "6029-ME", "BEUGRE", "Bassy Jean-Michel", "Agent de bureau", "Direction des Affaires Financières et du Patrimoine (DAFP)"),
    (163, "6030-ME", "YOUAN", "Lou Goure Hortense", "Agent de bureau", "Direction Régionale (DR) d'Agboville"),
    (164, "6031-ME", "COULIBALY", "Salimata Valerie", "Agent de bureau", "Direction Départementale (DD) de Bongouanou"),
    (165, "6032-ME", "GNACHOUE", "Ade Christine Zita", "Agent de bureau", "Direction Départementale (DD) de Bongouanou"),
    (166, "6033-ME", "RYAN", "Gneka", "Agent de bureau", "Direction départementale (DD) de Grand-Lahou"),
    (167, "6034-ME", "TRA", "Lou Gounan Eleonor Epse OUATTARA", "Agent de bureau", "Direction départementale (DD) de Danané"),
    (168, "6035-ME", "SAMAKE", "Mariam", "Agent de bureau", "Direction des Ressources Humaines (DRH)"),
    (169, "6036-ME", "N'GUESSAN", "Antoin Flore Liliane", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (170, "6037-ME", "BONI", "Chiakin Melaine Patricia", "Agent de bureau", "Direction Régionale (DR) d'Adzopé"),
    (171, "6038-ME", "ABOUA", "Prisca Melaine", "Agent de bureau", "Direction départementale (DD) d'Alépé"),
    (172, "6039-ME", "KOFFI", "Konan Sylvain", "Agent de bureau", "Direction du Domaine Public de l'Etat (DDPE)"),
    (173, "6040-ME", "BONZOU", "Kouamé Amakon Silva", "Agent de bureau", "Direction des Affaires Financières (DAF)"),
]


def create_excel():
    """Crée le fichier Excel avec tous les agents"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Agents AZING IVOIR"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Titre
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "LISTE DE PRÉSENCE DES AGENTS CONTRACTUELS - AZING IVOIR Sarl"
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    subtitle = ws["A2"]
    subtitle.value = "Ministère de l'Equipement et de l'Entretien Routier - Décembre 2025 - Effectif: 173 agents"
    subtitle.font = Font(bold=True, size=11, color="666666")
    subtitle.alignment = Alignment(horizontal="center")

    # En-têtes
    headers = ["N°", "MATRICULE", "NOM", "PRÉNOMS", "FONCTION", "STRUCTURE D'ACCUEIL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Données
    for agent in AGENTS:
        row = agent[0] + 4  # Décalage de 4 lignes (titre + sous-titre + vide + en-têtes)
        for col, value in enumerate(agent, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largeur des colonnes
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 55

    filepath = os.path.join(os.path.dirname(__file__), "Liste_Agents_AZING_IVOIR_Dec2025.xlsx")
    wb.save(filepath)
    print(f"Fichier Excel créé: {filepath}")
    print(f"Nombre d'agents: {len(AGENTS)}")
    return filepath


def import_to_database():
    """Importe les agents dans la base de données Django"""
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'empmanager.settings')
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))
    django.setup()

    from api.models import Employee, Department
    from datetime import date

    # Trouver ou créer le département AZING IVOIR
    dept, created = Department.objects.get_or_create(
        name="AZING IVOIR Sarl",
        defaults={
            "manager": "N'GUESSAN DAMBO Bénédicte",
            "description": "Prestation de service - Agents contractuels du Ministère de l'Equipement et de l'Entretien Routier"
        }
    )
    if created:
        print(f"Département créé: {dept.name}")
    else:
        print(f"Département existant: {dept.name}")

    created_count = 0
    skipped_count = 0

    for agent in AGENTS:
        num, matricule, nom, prenoms, fonction, direction = agent

        # Vérifier si l'agent existe déjà (par matricule)
        if Employee.objects.filter(matricule=matricule).exists():
            skipped_count += 1
            continue

        # Générer un email unique
        nom_clean = nom.lower().replace(' ', '').replace("'", '')
        prenom_clean = prenoms.lower().split()[0].replace("'", '')
        email_base = f"{nom_clean}.{prenom_clean}"
        email = f"{email_base}@azing-ivoir.ci"
        # Gérer les doublons d'email
        counter = 1
        while Employee.objects.filter(email=email).exists():
            email = f"{email_base}{counter}@azing-ivoir.ci"
            counter += 1

        Employee.objects.create(
            matricule=matricule,
            last_name=nom,
            first_name=prenoms,
            email=email,
            position=fonction,
            direction=direction,
            department=dept,
            hire_date=date(2025, 12, 1),
            status='active',
            salary=0,
        )
        created_count += 1

    print(f"\nRésultat de l'importation:")
    print(f"  - Agents créés: {created_count}")
    print(f"  - Agents ignorés (déjà existants): {skipped_count}")
    print(f"  - Total dans la base: {Employee.objects.filter(department=dept).count()}")


if __name__ == "__main__":
    # Étape 1: Créer le fichier Excel
    print("=" * 60)
    print("ÉTAPE 1: Création du fichier Excel")
    print("=" * 60)
    create_excel()

    # Étape 2: Importer dans la base de données
    print("\n" + "=" * 60)
    print("ÉTAPE 2: Importation dans la base de données Django")
    print("=" * 60)
    try:
        import_to_database()
    except Exception as e:
        print(f"\nErreur lors de l'importation en base: {e}")
        print("Le fichier Excel a été créé avec succès.")
        print("Vous pouvez importer les données plus tard en lançant ce script depuis le dossier backend.")
